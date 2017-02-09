import xlrd
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font

# xlsx format
border = Border(left=Side(border_style='medium',
                          color='FF000000'),
                right=Side(border_style='medium',
                           color='FF000000'),
                top=Side(border_style='medium',
                         color='FF000000'),
                bottom=Side(border_style='medium',
                            color='FF000000'),
                diagonal=Side(border_style='medium',
                              color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style='medium',
                             color='FF000000'),
                vertical=Side(border_style='medium',
                              color='FF000000'),
                horizontal=Side(border_style='medium',
                                color='FF000000')
                )

wb = load_workbook(r"C:\Users\Administrator\Desktop\报告单程序\报告单程序\模板.xlsx")
ws = wb.get_sheet_by_name('Sheet1')

# read names
arr = []
fisrt = r'C:\Users\Administrator\Desktop\报告单程序\报告单程序\1.xls'
book = xlrd.open_workbook(fisrt)
sheet = book.sheet_by_index(0)
nrows = sheet.nrows
nnames = nrows - 1

# print(nrows)
for n in range(1, nrows):
    arr.append(sheet.cell_value(n, 1))

# open files
a = r'C:\Users\Administrator\Desktop\报告单程序\报告单程序'
for i in range(1, 10):
    # print(i)
    b = a + r'\%s.xls' % i
    if os.path.exists(b):
        locals()['x%s' % i] = b
    else:
        m = i - 1
        break
# print(m)
sbarr = [1] * 200

for n1 in range(0, nnames):
    # read
    for n in range(1, m + 1):
        if n == 1:
            sbarr[0] = arr[n1]
        book = xlrd.open_workbook(locals()['x%s' % n])
        locals()['sheet%s' % n] = book.sheet_by_index(0)
        for n2 in range(1, 7):
            sbarr[6 * n - 6 + n2] = locals()['sheet%s' % n].cell_value(1 + n1, n2 + 1)

            # write
    ws.cell(row=3, column=2).value = sbarr[0]
    for n in range(1, 6 * m + 1):
        ny = (n - 1) // 6
        ns = n - ny * 6
        ws.cell(row=16 + ns, column=2 + ny).value = sbarr[n]
    for a in range(1, 54):
        for b in range(1, 16):
            ws.cell(row=a, column=b).border = border
    wb.save(r"C:\Users\Administrator\Desktop\报告单程序\报告单程序\生成\%s.xlsx" % sbarr[0])